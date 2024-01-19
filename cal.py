import pandas
import json
from icalendar import Calendar, Event
import datetime
import openpyxl
import tkinter as tk
from tkinter import filedialog

def is_green(hex_color):
    o = len(hex_color) - 6
    
    [r,g,b] = [int(hex_color[i:i+2],16) for i in range(0+o,6+o,2)]
    
    return g > r and g > b 

def main():
    # file dialog to get file path
    root = tk.Tk()
    root.withdraw()
    types = (("Excel files", "*.xlsx"), ("all files", "*.*"))   # file types
    path = filedialog.askopenfilename(filetypes=types)          # get file path

    colors = {}
    wb = openpyxl.load_workbook(path,data_only=True)
    fs = wb.active
    for row in range(1,fs.max_row+1):
        cell_color = fs.cell(column=8, row=row)
        bgColor = cell_color.fill.bgColor.index
        fgColor = cell_color.fill.fgColor.index

        colors[row] = fgColor

    excel_data = pandas.read_excel(path, sheet_name='Sheet1')
    formatted = json.loads(excel_data.to_json(orient='records'))

    # create calendar
    cal = Calendar()
    cal.add('prodid', '-//My calendar product//mxm.dk//') # set calendar product
    cal.add('version', '2.0') # set calendar version

    for i, each in enumerate(formatted):
        if     any([each[key] == None for key in each.keys()]) \
            or not is_green(colors[i + 2]): continue 

        dt_date         = datetime.datetime.fromtimestamp(each["Datum"] / 1000)
        dt_start_time   = datetime.datetime.strptime(each["Starttid"], "%H:%M:%S")
        dt_end_time     = datetime.datetime.strptime(each["Sluttid"], "%H:%M:%S")

        start           = datetime.datetime.combine(dt_date, dt_start_time.time())
        end             = datetime.datetime.combine(dt_date, dt_end_time.time())

        # create event
        event = Event()
        event.add('summary',    f"MOP: {each['Typ']}")         # set event title
        event.add('dtstart',    start)
        event.add('dtend',      end)
        event.add('dtstamp',    datetime.datetime.now())        # set event timestamp
        event.add('location',   each["Campus"])                 # set event location
        event.add('description', f"Added {datetime.datetime.now()}")      # set event description

        cal.add_component(event)

    # write to file
    newName = path.replace(".xlsx", ".ics")
    f = open(newName, 'wb')
    f.write(cal.to_ical())
    f.close()

    print(f"Done! Saved to {newName}")
    print(f"Added {len(cal.subcomponents)} events")
    
if __name__ == "__main__":
    main()